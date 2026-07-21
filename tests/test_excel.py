from pathlib import Path
from openpyxl import load_workbook
from core.models import Answer, Record
from core.excel.excel_writer import ExcelWriter
def test_writer_creates_required_sheets(tmp_path: Path):
    answers={f"Q{i}":Answer(f"Q{i}",answer_text="x",final_confidence=.9) for i in range(1,35)}
    path=ExcelWriter().write([Record("a","a.pdf",answers,.9,False)],tmp_path/"out.xlsx")
    assert set(load_workbook(path).sheetnames)=={"Responses","Review Queue","Audit","Summary","Processing Log"}
