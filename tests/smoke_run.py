"""Dependency-light smoke validation used when pytest is unavailable."""
from pathlib import Path
from openpyxl import load_workbook
from core.excel.excel_writer import ExcelWriter
from core.models import Answer, Record
from core.questionnaire.schema_loader import SchemaLoader

ROOT = Path(__file__).parents[1]
questions = SchemaLoader(ROOT / "config" / "questionnaire.json").load()
assert len(questions) == 34
answers = {f"Q{i}": Answer(f"Q{i}", answer_text="x", final_confidence=.9) for i in range(1, 35)}
target = ROOT / "tests" / "smoke.xlsx"
ExcelWriter().write([Record("smoke", "smoke.pdf", answers, .9, False)], target)
assert "Responses" in load_workbook(target).sheetnames
target.unlink()
print("smoke ok")
