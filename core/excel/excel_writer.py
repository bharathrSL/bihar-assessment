"""Single-workbook batch reporting with audit and review sheets."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from core.models import Answer, Record
from core.questionnaire.answer_mapper import AnswerMapper
class ExcelWriter:
    def write(self, records: list[Record], path: str | Path) -> Path:
        path=Path(path); path.parent.mkdir(parents=True,exist_ok=True); wb=Workbook(); responses=wb.active; responses.title="Responses"
        header=["Record ID","PDF","Confidence","Review"]+[f"Q{i}" for i in range(1,35)]; responses.append(header)
        for c in responses[1]: c.font=Font(bold=True); c.fill=PatternFill("solid",fgColor="D9EAF7")
        review=wb.create_sheet("Review Queue"); review.append(["Record ID","PDF","Question","AI Answer","Confidence","Reason"])
        audit=wb.create_sheet("Audit"); audit.append(["Record ID","Event"])
        summary=wb.create_sheet("Summary"); summary.append(["Metric","Value"])
        logs=wb.create_sheet("Processing Log"); logs.append(["Record ID","PDF","Confidence","Review"])
        for record in records:
            row = [
    record.record_id,
    record.pdf,
    record.confidence,
    record.review,
] + [
    AnswerMapper.value(record.answers.get(f"Q{i}", Answer(question_id=f"Q{i}")))
    for i in range(1, 35)
]
            responses.append(row); logs.append([record.record_id,record.pdf,record.confidence,record.review])
            for event in record.audit: audit.append([record.record_id,event])
            for answer in record.answers.values():
                if answer.review_required: review.append([record.record_id,record.pdf,answer.question_id,AnswerMapper.value(answer),answer.final_confidence,answer.raw_observations])
        summary.append(["Questionnaires",len(records)]); summary.append(["Review rate",sum(r.review for r in records)/len(records) if records else 0]); summary.append(["Average confidence",sum(r.confidence for r in records)/len(records) if records else 0])
        for ws in wb.worksheets: ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        wb.save(path); return path
